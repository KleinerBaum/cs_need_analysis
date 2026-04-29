from __future__ import annotations

import argparse
import hashlib
import json
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def _normalize_header(value: object) -> str:
    text = str(value or "").strip().casefold()
    text = re.sub(r"[^a-z0-9]+", "_", text).strip("_")
    return text


def _normalize_text(value: object) -> str:
    return str(value or "").strip()


def _parse_share_percent(value: object) -> float:
    if isinstance(value, (int, float)):
        return float(value)
    text = _normalize_text(value).replace(",", ".").replace("%", "")
    if not text:
        return 0.0
    return float(text)


def _derive_bucket(*, share_percent: float, fallback: str = "") -> str:
    normalized = fallback.strip().casefold()
    if normalized in {"must", "essential", "hasessentialskill"}:
        return "must"
    if normalized in {"nice", "optional", "hasoptionalskill"}:
        return "nice"
    return "must" if share_percent >= 50.0 else "nice"


def _slug(value: str) -> str:
    cleaned = re.sub(r"[^a-z0-9]+", "-", value.casefold()).strip("-")
    return cleaned or "unknown"


def _stable_skill_group_id(*, skill_group_uri: str, skill_group_label: str) -> str:
    seed = skill_group_uri or skill_group_label or "unknown"
    digest = hashlib.sha1(seed.encode("utf-8")).hexdigest()[:10]
    return f"skill-group-{_slug(skill_group_label or skill_group_uri)}-{digest}"


def _pick(row: dict[str, Any], *keys: str) -> str:
    for key in keys:
        value = _normalize_text(row.get(key))
        if value:
            return value
    return ""


def convert_xlsx_to_matrix_json(
    *,
    xlsx_path: Path,
    out_json_path: Path,
    version: str = "unknown",
    source: str | None = None,
) -> dict[str, Any]:
    workbook = load_workbook(filename=xlsx_path, data_only=True)
    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Workbook has no rows.")

    header_row = rows[0]
    headers = [_normalize_header(value) for value in header_row]
    if not any(headers):
        raise ValueError("Workbook header row is empty.")

    normalized_records: list[dict[str, Any]] = []
    for row_values in rows[1:]:
        if row_values is None:
            continue
        row = {headers[idx]: row_values[idx] for idx in range(min(len(headers), len(row_values)))}

        occupation_group = _pick(
            row,
            "occupation_group",
            "occupationgroup",
            "isco_group",
            "isco08",
            "isco_08",
            "isco08_code",
            "isco_code",
        )
        skill_group_uri = _pick(
            row,
            "skill_group_uri",
            "skillgroupuri",
            "skill_uri",
            "uri",
        )
        skill_group_label = _pick(
            row,
            "skill_group_label",
            "skillgrouplabel",
            "skill_title",
            "title",
            "label",
            "preferred_label",
            "preferredlabel",
            "skill_group",
        )
        if not occupation_group or not (skill_group_uri or skill_group_label):
            continue

        share_percent = _parse_share_percent(
            row.get("share_percent")
            or row.get("share")
            or row.get("share_pct")
            or row.get("percentage")
        )
        bucket = _derive_bucket(
            share_percent=share_percent,
            fallback=_pick(row, "bucket", "relation"),
        )
        skill_group_id = _pick(
            row,
            "skill_group_id",
            "skillgroupid",
            "group_id",
        ) or _stable_skill_group_id(
            skill_group_uri=skill_group_uri,
            skill_group_label=skill_group_label,
        )
        skill_uri = skill_group_uri or f"urn:esco:skill-group:{skill_group_id}"
        skill_title = skill_group_label or skill_group_id

        normalized_records.append(
            {
                "occupation_group": occupation_group,
                "skill_group_uri": skill_group_uri,
                "skill_group_id": skill_group_id,
                "skill_group_label": skill_group_label or skill_title,
                "share_percent": share_percent,
                "bucket": bucket,
                # Compatibility fields consumed by load_esco_matrix(...)
                "skill_uri": skill_uri,
                "skill_title": skill_title,
            }
        )

    payload = {
        "source": source or str(xlsx_path),
        "version": version,
        "records": normalized_records,
    }
    out_json_path.parent.mkdir(parents=True, exist_ok=True)
    out_json_path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return payload


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Convert ESCO Skills-Occupations Matrix XLSX to normalized JSON."
    )
    parser.add_argument("--xlsx", required=True, type=Path, help="Input XLSX path.")
    parser.add_argument("--out", required=True, type=Path, help="Output JSON path.")
    parser.add_argument("--version", default="unknown", help="Version metadata for output JSON.")
    parser.add_argument(
        "--source",
        default="",
        help="Optional source metadata (defaults to input XLSX path).",
    )
    args = parser.parse_args()
    convert_xlsx_to_matrix_json(
        xlsx_path=args.xlsx,
        out_json_path=args.out,
        version=str(args.version).strip() or "unknown",
        source=str(args.source).strip() or None,
    )


if __name__ == "__main__":
    main()
